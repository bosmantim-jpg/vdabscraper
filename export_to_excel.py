#!/usr/bin/env python3
"""Export scraped jobs from markdown files to Excel."""

import json
import re
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    exit(1)


def parse_markdown_job(file_path):
    """Parse a job markdown file and extract structured data."""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    job_data = {}

    # Extract title
    title_match = re.search(r'^# (.+)$', content, re.MULTILINE)
    job_data['title'] = title_match.group(1).strip() if title_match else '[MISSING]'

    # Extract fields from metadata lines
    bedrijf_match = re.search(r'^\*\*Bedrijf:\*\*\s*(.+?)$', content, re.MULTILINE)
    job_data['company'] = bedrijf_match.group(1).strip() if bedrijf_match else '[MISSING]'

    locatie_match = re.search(r'^\*\*Locatie:\*\*\s*(.+?)$', content, re.MULTILINE)
    job_data['location'] = locatie_match.group(1).strip() if locatie_match else '[MISSING]'

    contract_match = re.search(r'^\*\*Contract:\*\*\s*(.+?)$', content, re.MULTILINE)
    job_data['contract'] = contract_match.group(1).strip() if contract_match else '[MISSING]'

    salary_match = re.search(r'^\*\*Verloning:\*\*\s*(.+?)$', content, re.MULTILINE)
    job_data['salary'] = salary_match.group(1).strip() if salary_match else '[MISSING]'

    # Extract reference ID
    ref_match = re.search(r'^\*\*Referentie:\*\*\s*(.+?)$', content, re.MULTILINE)
    job_data['reference'] = ref_match.group(1).strip() if ref_match else '[MISSING]'

    # Extract URL
    url_match = re.search(r'^\*Bron:\s*(.+?)\s*\*$', content, re.MULTILINE)
    job_data['url'] = url_match.group(1).strip() if url_match else '[MISSING]'

    # Extract languages section
    languages = []
    lang_section_match = re.search(
        r'## Taalvereisten\s*\n\n(.+?)(?=\n##|$)',
        content,
        re.DOTALL
    )
    if lang_section_match:
        lang_section = lang_section_match.group(1)
        # Look for table rows
        table_rows = re.findall(r'\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|', lang_section)
        for row in table_rows:
            if row[0].strip() and row[0].strip() not in ['Taal', '---']:
                languages.append({
                    'language': row[0].strip(),
                    'level': row[1].strip(),
                    'required': 'Vereist' in row[2]
                })

    job_data['languages'] = '; '.join(
        f"{l['language']} ({l['level']}, {'Must' if l['required'] else 'Nice'})"
        for l in languages
    ) if languages else 'N/A'

    # Extract description (first 500 chars)
    desc_section_match = re.search(
        r'## Functiebeschrijving\s*\n\n(.+?)(?=\n---|\Z)',
        content,
        re.DOTALL
    )
    if desc_section_match:
        desc = desc_section_match.group(1).strip()
        job_data['description'] = desc[:500] + ('...' if len(desc) > 500 else '')
    else:
        job_data['description'] = '[MISSING]'

    return job_data


def export_to_excel():
    """Export all jobs from markdown files to Excel."""
    jobs_dir = Path('jobs')

    if not jobs_dir.exists():
        print("Error: jobs/ directory not found")
        return

    # Collect all jobs
    jobs = []
    md_files = sorted(jobs_dir.glob('*.md'))

    print(f"[*] Found {len(md_files)} job files")

    for file_path in md_files:
        try:
            job_data = parse_markdown_job(file_path)
            jobs.append(job_data)
        except Exception as e:
            print(f"[!] Error parsing {file_path}: {e}")

    if not jobs:
        print("No jobs to export")
        return

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 12

    # Create header row
    headers = ['Job Title', 'Company', 'Location', 'Contract Type', 'Salary', 'Languages', 'Description', 'Reference']
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Add job data
    for row_num, job in enumerate(jobs, 2):
        ws.cell(row=row_num, column=1).value = job.get('title', '')
        ws.cell(row=row_num, column=2).value = job.get('company', '')
        ws.cell(row=row_num, column=3).value = job.get('location', '')
        ws.cell(row=row_num, column=4).value = job.get('contract', '')
        ws.cell(row=row_num, column=5).value = job.get('salary', '')
        ws.cell(row=row_num, column=6).value = job.get('languages', '')
        ws.cell(row=row_num, column=7).value = job.get('description', '')
        ws.cell(row=row_num, column=8).value = job.get('reference', '')

        # Format data rows
        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'vdab_jobs_{timestamp}.xlsx'
    wb.save(output_file)

    print(f"[+] Export complete!")
    print(f"[+] File: {output_file}")
    print(f"[+] Jobs exported: {len(jobs)}")
    print(f"[+] Location: {Path(output_file).resolve()}")


if __name__ == '__main__':
    export_to_excel()
